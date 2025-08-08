"""
Pipelines: (1) WAR REPORT (daily XLSX -> consolidated analytics) and
           (2) CLIENT FUNDS LIFE CYCLE (monthly CSV -> multi-sheet Excel + CSVs)

Environment variables (overridable by flags):
  WAR_INPUT_DIR, WAR_OUTPUT_DIR, LIFE_INPUT_CSV, LIFE_OUTPUT_DIR
"""

from __future__ import annotations

import os
import sys
import argparse
import logging
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# ───────────────────────────── logging ─────────────────────────────

def _setup_logging(level: str = "INFO") -> None:
    lvl = getattr(logging, level.upper(), logging.INFO)
    logging.basicConfig(
        level=lvl,
        format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
        stream=sys.stdout,
    )


log = logging.getLogger("pipelines")


# ───────────────────────────── helpers ─────────────────────────────

def ensure_dir(path: str | Path) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def write_sheet(df: pd.DataFrame, excel_path: Path, sheet_name: str, index: bool = False) -> None:
    """
    Append/replace a worksheet in an Excel file safely.
    - Creates the file if it doesn't exist
    - If the sheet exists, removes and rewrites it
    """
    sheet_name = (sheet_name or "Sheet")[:31]  # Excel limit

    if excel_path.exists():
        # load existing workbook, then re-open writer in append mode with the same book
        book = load_workbook(excel_path)
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a") as writer:
            writer.book = book
            # make existing sheets visible to pandas (avoids duplicate-sheet glitches)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            if sheet_name in writer.book.sheetnames:
                std = writer.book[sheet_name]
                writer.book.remove(std)
            df.to_excel(writer, sheet_name=sheet_name, index=index)
    else:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index)


def _coerce_numeric(s: pd.Series) -> pd.Series:
    """Robust numeric coercion that tolerates commas-as-decimal and stray spaces."""
    if s.dtype.kind in "biufc":
        return s
    return pd.to_numeric(
        s.astype(str).str.replace(",", ".", regex=False).str.replace(" ", "", regex=False),
        errors="coerce",
    )


# ─────────────────────── 1) WAR REPORT PIPELINE ───────────────────────

@dataclass
class WarReportConfig:
    input_dir: str
    output_dir: str
    default_min_date: datetime = datetime(2024, 1, 2)
    default_end_date: datetime = datetime.today()
    excel_generic_name: str = "Dados_Continuos.xlsx"  # used when full/continuous period is selected
    file_glob: str = "*.xlsx"
    read_usecols: str = "A:I"
    read_skiprows: int = 10


def build_consolidated_df(cfg: WarReportConfig, start_dt: datetime, end_dt: datetime) -> pd.DataFrame:
    """
    Read XLSX files in cfg.input_dir, parse block (usecols, skiprows), coerce date column,
    filter by [start_dt, end_dt], and return cleaned, concatenated DataFrame with unified names:
      - 'Fundo' (from 'NOMBRE'), 'Data' (from 'Fecha Operación')
    """
    input_dir = Path(cfg.input_dir)
    files = sorted(input_dir.glob(cfg.file_glob))
    if not files:
        log.warning("No files matching %s in %s", cfg.file_glob, cfg.input_dir)

    all_rows: List[pd.DataFrame] = []
    for f in files:
        try:
            df = pd.read_excel(f, usecols=cfg.read_usecols, skiprows=cfg.read_skiprows)
        except Exception as e:
            log.warning("Skipping %s: %s", f.name, e)
            continue

        # Map common column variants
        colmap: Dict[str, str] = {}
        for c in df.columns:
            c_norm = str(c).strip()
            lc = c_norm.lower()
            if lc.startswith("fecha"):
                colmap[c] = "Fecha Operación"
            elif lc in {"nombre", "nombre ", "nome", "nomb", "nombre do fundo"}:
                colmap[c] = "NOMBRE"
            else:
                colmap[c] = c_norm
        df = df.rename(columns=colmap)

        if "Fecha Operación" not in df.columns:
            log.warning("File %s missing 'Fecha Operación' column. Skipping.", f.name)
            continue

        # Date coercion: accepts 'YYYYMMDD' or other excel datetime
        df["Fecha Operación"] = pd.to_datetime(df["Fecha Operación"], errors="coerce", format="%Y%m%d").fillna(
            pd.to_datetime(df["Fecha Operación"], errors="coerce")
        )
        df = df.dropna(subset=["Fecha Operación"])
        mask = (df["Fecha Operación"] >= start_dt) & (df["Fecha Operación"] <= end_dt)
        df = df.loc[mask].copy()
        if df.empty:
            continue

        # Normalize likely numeric columns if present
        for c in ["Patrimonio", "COM", "VEN", "COM > VEN", "COM < VEN", "%"]:
            if c in df.columns:
                df.loc[:, c] = _coerce_numeric(df[c])

        all_rows.append(df)

    if not all_rows:
        return pd.DataFrame(columns=["FUM", "Fecha Operación", "NOMBRE", "Patrimonio", "COM", "VEN", "COM > VEN", "COM < VEN", "%"])

    dados = pd.concat(all_rows, ignore_index=True)

    # Drop duplicates on reasonable subset
    base_subset = [c for c in ["FUM", "Fecha Operación", "NOMBRE"] if c in dados.columns]
    if base_subset:
        dados = dados.drop_duplicates(subset=base_subset, keep="last")

    # Unified names
    if "NOMBRE" in dados.columns:
        dados = dados.rename(columns={"NOMBRE": "Fundo"})
    if "Fecha Operación" in dados.columns:
        dados = dados.rename(columns={"Fecha Operación": "Data"})

    sort_cols = [c for c in ["FUM", "Data"] if c in dados.columns]
    if sort_cols:
        dados = dados.sort_values(by=sort_cols)
    dados = dados.reset_index(drop=True)

    return dados


# ────────────────────────── Outliers ──────────────────────────

def outliers_by_fund(df: pd.DataFrame, col: str = "%", k: float = 2.0) -> pd.DataFrame:
    if ("Fundo" not in df.columns) or (col not in df.columns):
        return pd.DataFrame()

    out_all: List[pd.DataFrame] = []
    for fundo, g in df.groupby("Fundo", sort=False):
        series = g[col].dropna()
        if series.empty:
            continue
        mean, std = series.mean(), series.std()
        if pd.isna(std) or std == 0:
            continue
        low, high = mean - k * std, mean + k * std
        sel = g[(g[col] < low) | (g[col] > high)].copy()
        if not sel.empty:
            sel.loc[:, "Fundo"] = fundo
            sel.loc[:, "Limite_Baixo"] = low
            sel.loc[:, "Limite_Alto"] = high
            out_all.append(sel)

    if not out_all:
        return pd.DataFrame(columns=list(df.columns) + ["Fundo", "Limite_Baixo", "Limite_Alto"])
    return pd.concat(out_all, ignore_index=True)


def outliers_by_day(df: pd.DataFrame, col: str = "%", k: float = 2.0) -> pd.DataFrame:
    if ("Data" not in df.columns) or (col not in df.columns):
        return pd.DataFrame()

    out_all: List[pd.DataFrame] = []
    for dt, g in df.groupby("Data", sort=False):
        series = g[col].dropna()
        if series.empty:
            continue
        mean, std = series.mean(), series.std()
        if pd.isna(std) or std == 0:
            continue
        low, high = mean - k * std, mean + k * std
        sel = g[(g[col] < low) | (g[col] > high)].copy()
        if not sel.empty:
            sel.loc[:, "Data"] = pd.to_datetime(dt)
            sel.loc[:, "Limite_Baixo"] = low
            sel.loc[:, "Limite_Alto"] = high
            out_all.append(sel)

    if not out_all:
        return pd.DataFrame(columns=list(df.columns) + ["Data", "Limite_Baixo", "Limite_Alto"])
    return pd.concat(out_all, ignore_index=True)


# ───────────────────── Stats & Rankings ─────────────────────

def std_positive(x: pd.Series) -> float:
    return x[x > 0].std()


def std_negative(x: pd.Series) -> float:
    return x[x < 0].std()


def compute_stats_and_rankings(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    needed_cols = ["Fundo", "Patrimonio", "COM", "VEN", "COM > VEN", "COM < VEN", "%"]
    missing = [c for c in needed_cols if c not in df.columns]

    if missing:
        agg: Dict[str, List] = {}
        if "Patrimonio" in df.columns:
            agg["Patrimonio"] = ["mean", "std", "min", "max"]
        if "COM" in df.columns:
            agg["COM"] = ["sum", "mean", "std"]
        if "VEN" in df.columns:
            agg["VEN"] = ["sum", "mean", "std"]
        if "COM > VEN" in df.columns:
            agg["COM > VEN"] = ["mean"]
        if "COM < VEN" in df.columns:
            agg["COM < VEN"] = ["mean"]
        if "%" in df.columns:
            agg["%"] = ["mean", "std", "min", "max", std_positive, std_negative]
    else:
        agg = {
            "Patrimonio": ["mean", "std", "min", "max"],
            "COM": ["sum", "mean", "std"],
            "VEN": ["sum", "mean", "std"],
            "COM > VEN": ["mean"],
            "COM < VEN": ["mean"],
            "%": ["mean", "std", "min", "max", std_positive, std_negative],
        }

    if "Fundo" not in df.columns or not agg:
        return pd.DataFrame(), pd.DataFrame()

    stats = df.groupby("Fundo").agg(agg).fillna(0)
    stats.columns = ["_".join([c for c in t if c]).strip() for t in stats.columns.to_flat_index()]

    r = pd.DataFrame(index=stats.index)
    if "Patrimonio_mean" in stats.columns:
        r["Patrimonio Mean"] = stats["Patrimonio_mean"].rank(ascending=False, method="dense")
    if "COM_sum" in stats.columns:
        r["Total COM"] = stats["COM_sum"].rank(ascending=False, method="dense")
    if "VEN_sum" in stats.columns:
        r["Total VEN"] = stats["VEN_sum"].rank(ascending=False, method="dense")
    if "COM > VEN_mean" in stats.columns:
        r["Avg (COM > VEN)"] = stats["COM > VEN_mean"].rank(ascending=False, method="dense")
    if "COM < VEN_mean" in stats.columns:
        r["Avg (COM < VEN)"] = stats["COM < VEN_mean"].rank(ascending=True, method="dense")
    if "%_mean" in stats.columns:
        r["Avg %"] = stats["%_mean"].rank(ascending=False, method="dense")
    if "%_std_positive" in stats.columns:
        r["Std % (Positive)"] = stats["%_std_positive"].rank(ascending=False, method="dense")
    if "%_std_negative" in stats.columns:
        r["Std % (Negative)"] = stats["%_std_negative"].rank(ascending=True, method="dense")

    # Stable sort by first column if exists
    if not r.empty:
        first_col = r.columns[0]
        r = r.sort_values(by=first_col)
    return stats, r


# ─────────────────────────── Correlations ───────────────────────────

def correlation_tables(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Builds:
      - full correlation matrix of 'Patrimonio' % change per fund
      - pair list (lower triangle) sorted
      - per-fund best positive and best negative partner
      - top table with rank & label
    """
    required = {"Fundo", "Data", "Patrimonio"}
    if not required.issubset(df.columns):
        return {
            "Matriz_Correlacao_Completa": pd.DataFrame(),
            "Correlacoes_Completas": pd.DataFrame(),
            "Resumo_por_Fundo": pd.DataFrame(),
            "Top_Correlacoes": pd.DataFrame(),
        }

    work = df[["Fundo", "Data", "Patrimonio"]].dropna().copy()
    work["Data"] = pd.to_datetime(work["Data"])
    work = work.sort_values(["Fundo", "Data"])
    work["Pct_Change"] = work.groupby("Fundo")["Patrimonio"].pct_change()

    pivot = work.pivot_table(index="Data", columns="Fundo", values="Pct_Change")
    corr = pivot.corr(min_periods=2)

    # Pairwise (lower triangle only)
    mask_lower = np.tril(np.ones(corr.shape), k=-1).astype(bool)
    pairs = corr.where(mask_lower).stack().reset_index()
    pairs.columns = ["Fund 1", "Fund 2", "Correlation"]
    pairs = pairs.sort_values(by="Correlation", ascending=False, kind="mergesort")

    # Per-fund best +/- partner
    best_pos, best_neg = [], []
    for f in corr.columns:
        s = corr[f].drop(labels=[f]).dropna()
        if s.empty:
            best_pos.append((f, None, np.nan))
            best_neg.append((f, None, np.nan))
        else:
            ppos = s.idxmax()
            pneg = s.idxmin()
            best_pos.append((f, ppos, s.loc[ppos]))
            best_neg.append((f, pneg, s.loc[pneg]))

    resumo = pd.DataFrame(
        {
            "Fund": [t[0] for t in best_pos],
            "Best_Positive_Partner": [t[1] for t in best_pos],
            "Best_Positive_Corr": [t[2] for t in best_pos],
            "Best_Negative_Partner": [t[1] for t in best_neg],
            "Best_Negative_Corr": [t[2] for t in best_neg],
        }
    )
    resumo = resumo.merge(corr.abs().mean().rename("Avg_Abs_Corr"), left_on="Fund", right_index=True, how="left")
    resumo = resumo.merge(corr.apply(lambda x: x[x > 0].mean(), axis=1).rename("Avg_Pos_Corr"), left_on="Fund", right_index=True, how="left")

    top = pairs.copy()
    top["Rank"] = range(1, len(top) + 1)

    def _fmt(row) -> str:
        try:
            return f"Top {row['Rank']}: {row['Fund 1']} × {row['Fund 2']} (corr={row['Correlation']:.2f})"
        except Exception:
            return f"Top {row['Rank']}: {row['Fund 1']} × {row['Fund 2']} (corr=NA)"

    top["Label"] = top.apply(_fmt, axis=1)

    return {
        "Matriz_Correlacao_Completa": corr,
        "Correlacoes_Completas": pairs,
        "Resumo_por_Fundo": resumo,
        "Top_Correlacoes": top[["Rank", "Fund 1", "Fund 2", "Correlation", "Label"]],
    }


# ───────────────────────── Optional exports ─────────────────────────

def export_fund_evolution(df: pd.DataFrame, fund: str, excel_path: Path) -> None:
    base_cols = ["Data", "Fundo"]
    extras = [c for c in ["Patrimonio", "COM", "VEN", "%", "COM > VEN", "COM < VEN"] if c in df.columns]
    cols = base_cols + extras
    ser = df.loc[df["Fundo"] == fund, cols].copy().sort_values("Data")
    if ser.empty:
        log.info("[Evolution] No rows for '%s'.", fund)
        return
    sheet = f"Evolution_{fund}"[:31]
    write_sheet(ser, excel_path, sheet, index=False)
    log.info("[Evolution] Sheet '%s' written (%d rows).", sheet, len(ser))


def export_fund_arima(df: pd.DataFrame, fund: str, horizon: int, excel_path: Path) -> None:
    try:
        from statsmodels.tsa.statespace.sarimax import SARIMAX  # type: ignore
    except Exception as e:
        log.warning("[ARIMA] statsmodels not available: %s", e)
        return

    ser = df.loc[df["Fundo"] == fund, ["Data", "Patrimonio"]].dropna().copy().sort_values("Data")
    if len(ser) < 10:
        log.info("[ARIMA] Too few points for '%s' (%d). Skipping.", fund, len(ser))
        return

    ts = ser.set_index("Data").asfreq("B")
    ts["Patrimonio"] = ts["Patrimonio"].interpolate("time")

    try:
        model = SARIMAX(ts["Patrimonio"], order=(1, 1, 1), enforce_stationarity=False, enforce_invertibility=False)
        res = model.fit(disp=False)
        fc = res.get_forecast(steps=horizon)
        pred = fc.predicted_mean.rename("Forecast")
        conf = fc.conf_int(alpha=0.05)
        if conf.shape[1] == 2:
            conf.columns = ["Low", "High"]
        else:
            conf = conf.rename(columns=lambda c: "Low" if "lower" in c.lower() else ("High" if "upper" in c.lower() else c))
        out = pd.concat([pred, conf], axis=1).reset_index().rename(columns={"index": "Data"})
        out.insert(0, "Fundo", fund)

        sheet = f"Forecast_{fund}"[:31]
        write_sheet(out, excel_path, sheet, index=False)
        log.info("[ARIMA] Sheet '%s' written (%d rows).", sheet, len(out))
    except Exception as e:
        log.warning("[ARIMA] Error for '%s': %s", fund, e)


def run_war_report(
    input_dir: str,
    output_dir: str,
    start: Optional[str] = None,
    end: Optional[str] = None,
    excel_generic_name: str = "Dados_Continuos.xlsx",
    file_glob: str = "*.xlsx",
    read_usecols: str = "A:I",
    read_skiprows: int = 10,
    evolution_funds: Optional[List[str]] = None,
    arima_funds: Optional[List[str]] = None,
    arima_horizon: int = 10,
) -> Path:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    cfg = WarReportConfig(
        input_dir=input_dir,
        output_dir=output_dir,
        excel_generic_name=excel_generic_name,
        file_glob=file_glob,
        read_usecols=read_usecols,
        read_skiprows=read_skiprows,
    )
    out_dir = ensure_dir(cfg.output_dir)

    # Dates
    start_dt = datetime.strptime(start, "%Y-%m-%d") if start else cfg.default_min_date
    end_dt = datetime.strptime(end, "%Y-%m-%d") if end else cfg.default_end_date
    log.info("Analysis period: %s to %s", start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d"))

    # Build consolidated
    dados = build_consolidated_df(cfg, start_dt, end_dt)
    if dados.empty:
        log.warning("No data found for the selected period.")
        return out_dir / "NO_DATA.xlsx"

    # Decide output filename
    data_real_inicio = pd.to_datetime(dados["Data"]).min()
    data_real_fim = pd.to_datetime(dados["Data"]).max()
    continuous = (start_dt.date() == cfg.default_min_date.date()) and (end_dt.date() >= data_real_fim.date())
    excel_name = cfg.excel_generic_name if continuous else f"Dados_Filtrados_{data_real_inicio:%Y-%m-%d}_ate_{data_real_fim:%Y-%m-%d}.xlsx"
    excel_path = out_dir / excel_name

    # Base sheet
    write_sheet(dados, excel_path, "Dados_Concatenados", index=False)

    # Outliers
    out_fund = outliers_by_fund(dados, col="%", k=2)
    out_day = outliers_by_day(dados, col="%", k=2)
    if not out_fund.empty:
        write_sheet(out_fund, excel_path, "Outliers_Fundo", index=False)
    if not out_day.empty:
        write_sheet(out_day, excel_path, "Outliers_Dia", index=False)

    # Stats & Rankings
    stats, ranks = compute_stats_and_rankings(dados)
    if not stats.empty:
        write_sheet(stats, excel_path, "Estatisticas", index=True)
    if not ranks.empty:
        write_sheet(ranks, excel_path, "Rankings", index=True)

    # Correlations
    ct = correlation_tables(dados)
    for sheet, df in ct.items():
        if not df.empty:
            write_sheet(df, excel_path, sheet, index=(sheet == "Matriz_Correlacao_Completa"))

    # Optional: per-fund evolution & ARIMA
    for f in (evolution_funds or []):
        export_fund_evolution(dados, f, excel_path)
    for f in (arima_funds or []):
        export_fund_arima(dados, f, arima_horizon, excel_path)

    log.info("✅ WAR REPORT complete. Excel written at: %s", excel_path)
    return excel_path


# ───────────────────── 2) LIFE CYCLE PIPELINE ─────────────────────

@dataclass
class LifeCfg:
    input_csv: str
    out_dir: str
    excel_name: str = "Funds_Life.xlsx"
    csv_subdir: str = "csv"
    first_month_known: str = "2023-11-30"  # first snapshot (special rule)


def ensure_dirs_life(cfg: LifeCfg) -> Tuple[Path, Path]:
    out_dir = ensure_dir(cfg.out_dir)
    csv_dir = ensure_dir(out_dir / cfg.csv_subdir)
    return out_dir, csv_dir


def save_all_outputs(dfs: Dict[str, pd.DataFrame], out_excel: Path, out_csv_dir: Path) -> None:
    # Write Excel
    with pd.ExcelWriter(out_excel, engine="openpyxl", datetime_format="yyyy-mm-dd", date_format="yyyy-mm-dd") as writer:
        for sheet, df in dfs.items():
            df2 = df.copy()
            for col in df2.columns:
                if pd.api.types.is_period_dtype(df2[col]):
                    df2.loc[:, col] = df2[col].astype(str)
            df2.to_excel(writer, sheet_name=(sheet[:31] or "Sheet"), index=False)

    # Write CSVs (Power BI friendly)
    for sheet, df in dfs.items():
        df2 = df.copy()
        for col in df2.columns:
            if pd.api.types.is_period_dtype(df2[col]):
                df2.loc[:, col] = df2[col].astype(str)
        df2.to_csv(out_csv_dir / f"{sheet}.csv", index=False, encoding="utf-8-sig")


# ───────────────────────── Load + Prep ─────────────────────────

def load_and_prepare(cfg: LifeCfg) -> pd.DataFrame:
    df = pd.read_csv(cfg.input_csv, delimiter=",")
    df["fecha_contenido"] = pd.to_datetime(df["fecha_contenido"], format="%Y%m%d", errors="coerce")
    df["des_produto"] = df["des_produto"].astype(str).str.strip()
    df["nr_unidades"] = _coerce_numeric(df["nr_unidades"])
    df["nuc"] = pd.to_numeric(df["nuc"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["fecha_contenido", "des_produto", "nr_unidades", "nuc"])

    # Aggregate duplicates (same nuc, fund, date)
    df = df.groupby(["nuc", "des_produto", "fecha_contenido"], as_index=False)["nr_unidades"].sum()
    df["mes_ano"] = df["fecha_contenido"].dt.to_period("M")
    return df


# ─────────────────────────── Tagging ───────────────────────────

def tag_movements(df: pd.DataFrame, cfg: LifeCfg) -> pd.DataFrame:
    df = df.sort_values(["nuc", "des_produto", "fecha_contenido"]).copy()
    df["prev_nr_unidades"] = df.groupby(["nuc", "des_produto"])["nr_unidades"].shift(1)
    df["next_nr_unidades"] = df.groupby(["nuc", "des_produto"])["nr_unidades"].shift(-1)

    first_cutoff = pd.Timestamp(cfg.first_month_known)

    def classify(row) -> str:
        prev = row["prev_nr_unidades"]
        curr = row["nr_unidades"]
        nxt = row["next_nr_unidades"]
        dt = row["fecha_contenido"]

        if pd.isna(prev):
            return "buy" if dt > first_cutoff else "check_november"
        if curr > prev:
            return "increase"
        if curr < prev:
            if (curr == 0) or pd.isna(nxt):
                return "sell"
            return "reduce"
        return "keep"

    df["tag"] = df.apply(classify, axis=1)

    # Resolve 'check_november' by looking ahead for activity
    nov_mask = df["tag"].eq("check_november")
    if nov_mask.any():
        for idx in df.index[nov_mask]:
            nuc = df.at[idx, "nuc"]
            prod = df.at[idx, "des_produto"]
            dt = df.at[idx, "fecha_contenido"]
            subsequent = df[(df["nuc"] == nuc) & (df["des_produto"] == prod) & (df["fecha_contenido"] > dt)]
            df.at[idx, "tag"] = "buy" if not subsequent.empty else "sell"

    return df


# ─────────────────────────── Metrics ───────────────────────────

def kpis_funds_per_client(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    dist = df.groupby("nuc")["des_produto"].nunique().rename("num_funds").reset_index()
    kpi = pd.DataFrame(
        {
            "metric": ["mean", "median", "mode", "max", "min"],
            "value": [
                dist["num_funds"].mean(),
                dist["num_funds"].median(),
                dist["num_funds"].mode().iloc[0] if not dist["num_funds"].mode().empty else np.nan,
                dist["num_funds"].max(),
                dist["num_funds"].min(),
            ],
        }
    )
    max_funds = dist["num_funds"].max()
    clients_max = dist.loc[dist["num_funds"] == max_funds, ["nuc", "num_funds"]].reset_index(drop=True)
    return dist, kpi, clients_max


def clients_per_fund(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    total = df.groupby("des_produto")["nuc"].nunique().rename("unique_clients").reset_index()
    per_month = df.groupby([df["mes_ano"], "des_produto"])["nuc"].nunique().rename("unique_clients").reset_index()
    return total, per_month


def bpi_stats(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    out: Dict[str, pd.DataFrame] = {}
    bpi = df[df["des_produto"].str.contains("BPI", case=False, na=False)].copy()
    if bpi.empty:
        out["BPI_Stats"] = pd.DataFrame(columns=["stat", "value"])
        out["BPI_Monthly_Volume"] = pd.DataFrame(columns=["fecha_contenido", "nr_unidades"])
        out["BPI_Extra"] = pd.DataFrame(columns=["stat", "nr_unidades", "date"])
        return out

    desc = bpi["nr_unidades"].describe().to_frame("value").reset_index().rename(columns={"index": "stat"})
    mode_series = bpi["nr_unidades"].mode()
    mode_val = mode_series.iloc[0] if not mode_series.empty else np.nan
    by_month = bpi.groupby("fecha_contenido", as_index=False)["nr_unidades"].sum().sort_values("fecha_contenido")

    if not by_month.empty:
        date_max = by_month.loc[by_month["nr_unidades"].idxmax(), "fecha_contenido"]
        max_val = by_month["nr_unidades"].max()
        date_min = by_month.loc[by_month["nr_unidades"].idxmin(), "fecha_contenido"]
        min_val = by_month["nr_unidades"].min()
        extra = pd.DataFrame(
            [
                {"stat": "mode", "nr_unidades": mode_val, "date": pd.NaT},
                {"stat": "date_max_volume", "nr_unidades": max_val, "date": date_max},
                {"stat": "date_min_volume", "nr_unidades": min_val, "date": date_min},
            ]
        )
    else:
        extra = pd.DataFrame([{"stat": "mode", "nr_unidades": mode_val, "date": pd.NaT}])

    out["BPI_Stats"] = desc
    out["BPI_Monthly_Volume"] = by_month
    out["BPI_Extra"] = extra
    return out


def combinations(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    base = df.groupby("nuc")["des_produto"].agg(lambda x: set(x)).reset_index()
    base["num_funds"] = base["des_produto"].apply(len)
    multi = base[base["num_funds"] > 1].copy()
    multi["funds_combo"] = multi["des_produto"].apply(lambda s: ", ".join(sorted(s)))
    comb_freq = (
        multi.groupby("funds_combo").size().reset_index(name="count").sort_values("count", ascending=False, kind="mergesort")
    )
    comb_freq["num_funds"] = comb_freq["funds_combo"].apply(lambda x: x.count(",") + 1)
    resume = pd.DataFrame([{"total_unique_combinations": comb_freq.shape[0]}])
    return comb_freq, resume


def churn_metrics(df_tagged: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    sells = df_tagged[df_tagged["tag"].eq("sell")][["nuc", "fecha_contenido"]].copy()

    def no_reinvest(row) -> bool:
        fut = df_tagged[(df_tagged["nuc"] == row["nuc"]) & (df_tagged["fecha_contenido"] > row["fecha_contenido"])]
        return fut.empty

    if not sells.empty:
        sells.loc[:, "no_reinvestment"] = sells.apply(no_reinvest, axis=1)
        churn_candidates = sells[sells["no_reinvestment"]].copy()
        churn_by_month = (
            churn_candidates.groupby(churn_candidates["fecha_contenido"].dt.to_period("M"))["nuc"]
            .nunique()
            .rename("clients_churn")
            .reset_index()
        )
    else:
        churn_by_month = pd.DataFrame(columns=["fecha_contenido", "clients_churn"])

    active_by_month = (
        df_tagged.groupby(df_tagged["fecha_contenido"].dt.to_period("M"))["nuc"].nunique().rename("active_clients").reset_index()
    )

    churn_global = pd.merge(active_by_month, churn_by_month, on="fecha_contenido", how="left").fillna({"clients_churn": 0})
    churn_global["churn_rate_%"] = (churn_global["clients_churn"] / churn_global["active_clients"]) * 100

    # churn per fund per month
    df_ = df_tagged.copy()
    df_["mes_ano"] = df_["fecha_contenido"].dt.to_period("M")
    holders = df_.groupby(["mes_ano", "des_produto"])["nuc"].nunique().rename("clients_with_fund").reset_index()

    df_sort = df_.sort_values(["nuc", "des_produto", "fecha_contenido"]).copy()
    df_sort["next_month_for_pair"] = df_sort.groupby(["nuc", "des_produto"])["mes_ano"].shift(-1)
    churn_pairs = df_sort[df_sort["next_month_for_pair"].isna()][["nuc", "des_produto", "mes_ano"]]
    churn_pairs = churn_pairs.groupby(["mes_ano", "des_produto"])["nuc"].nunique().rename("clients_churn_fund").reset_index()

    churn_fund_month = pd.merge(holders, churn_pairs, on=["mes_ano", "des_produto"], how="left").fillna({"clients_churn_fund": 0})
    churn_fund_month["churn_rate_fund_%"] = (churn_fund_month["clients_churn_fund"] / churn_fund_month["clients_with_fund"]) * 100
    return churn_global, churn_fund_month


def acquisition_and_retention(df_tagged: pd.DataFrame):
    buys = df_tagged[df_tagged["tag"].eq("buy")].copy()
    acq_global = (
        buys.groupby(buys["fecha_contenido"].dt.to_period("M"))["nuc"].nunique().rename("clients_buy").reset_index()
    )
    acq_fund = (
        buys.groupby([buys["fecha_contenido"].dt.to_period("M"), "des_produto"])["nuc"]
        .nunique()
        .rename("clients_buy")
        .reset_index()
        .rename(columns={"fecha_contenido": "mes_ano"})
    )
    top_acq_abs = (
        buys.groupby("des_produto")["nuc"].nunique().rename("unique_buyers").reset_index().sort_values("unique_buyers", ascending=False)
    )

    df_ = df_tagged.sort_values(["nuc", "des_produto", "fecha_contenido"]).copy()
    df_["mes_ano"] = df_["fecha_contenido"].dt.to_period("M")
    df_["next_month"] = df_.groupby(["nuc", "des_produto"])["mes_ano"].shift(-1)

    base_ret = df_.groupby(["mes_ano", "des_produto"])["nuc"].nunique().rename("clients_t").reset_index()
    cont_ret = (
        df_[~df_["next_month"].isna()].groupby(["mes_ano", "des_produto"])["nuc"].nunique().rename("clients_t_continue").reset_index()
    )
    ret = pd.merge(base_ret, cont_ret, on=["mes_ano", "des_produto"], how="left").fillna({"clients_t_continue": 0})
    ret["retention_rate_%"] = (ret["clients_t_continue"] / ret["clients_t"]) * 100

    return acq_global, acq_fund, top_acq_abs, ret


def client_profiles(df_tagged: pd.DataFrame) -> pd.DataFrame:
    df_ = df_tagged.copy()
    df_["mes_ano"] = df_["fecha_contenido"].dt.to_period("M")
    max_month = df_["mes_ano"].max()

    quantity = df_.groupby("nuc")["des_produto"].nunique().rename("quantity")
    freq = (
        df_[df_["tag"].isin(["buy", "sell", "increase", "reduce"])].groupby("nuc").size().rename("frequency")
    )
    last_month = df_.groupby("nuc")["mes_ano"].max().rename("last_month")
    recency = (max_month - last_month).apply(lambda x: x.n).rename("recency_months")

    last_move = (
        df_[df_["tag"].isin(["buy", "sell", "increase", "reduce"])].groupby("nuc")["mes_ano"].max().rename("last_move_month")
    )
    inactivity = (max_month - last_move).apply(lambda x: x.n if pd.notna(x) else np.nan).rename("inactivity_months")

    has_current = df_[df_["mes_ano"].eq(max_month)].groupby("nuc").size().rename("has_current")
    is_current = has_current.gt(0).rename("is_current_client")

    months_since_churn = recency.where(~is_current, other=0).rename("months_since_churn")

    profiles = pd.concat([quantity, freq, recency, inactivity, is_current, months_since_churn], axis=1).fillna(
        {"frequency": 0, "inactivity_months": np.nan}
    )
    return profiles.reset_index()


def flows_transitions(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df_ = df.copy()
    df_["mes_ano"] = df_["fecha_contenido"].dt.to_period("M")
    holds = df_.groupby(["nuc", "mes_ano"])["des_produto"].apply(lambda s: set(s)).reset_index(name="funds")

    rows: List[Dict[str, object]] = []
    for nuc, grp in holds.groupby("nuc", sort=False):
        grp = grp.sort_values("mes_ano")
        prev_set: Optional[set] = None
        prev_month = None
        for _, row in grp.iterrows():
            month = row["mes_ano"]
            curr_set = row["funds"]
            if prev_set is not None:
                exited = list(prev_set - curr_set)
                entered = list(curr_set - prev_set)
                if len(exited) == 1 and len(entered) == 1:
                    rows.append({"nuc": nuc, "from_fund": exited[0], "to_fund": entered[0], "from_month": prev_month, "to_month": month})
                else:
                    for a in exited:
                        rows.append({"nuc": nuc, "from_fund": a, "to_fund": "(OUT)", "from_month": prev_month, "to_month": month})
                    for b in entered:
                        rows.append({"nuc": nuc, "from_fund": "(IN)", "to_fund": b, "from_month": prev_month, "to_month": month})
            prev_set, prev_month = curr_set, month

    detail = pd.DataFrame(rows)
    if detail.empty:
        matrix = pd.DataFrame(columns=["from_fund", "to_fund", "count"])
    else:
        matrix = (
            detail.groupby(["from_fund", "to_fund"]).size().rename("count").reset_index().sort_values("count", ascending=False, kind="mergesort")
        )
    return detail, matrix


def run_life_cycle(
    csv_path: str,
    output_dir: str,
    excel_name: str = "Funds_Life.xlsx",
    first_month_known: str = "2023-11-30",
) -> Tuple[Path, Path]:
    cfg = LifeCfg(input_csv=csv_path, out_dir=output_dir, excel_name=excel_name, first_month_known=first_month_known)
    out_dir, csv_dir = ensure_dirs_life(cfg)
    out_excel = out_dir / cfg.excel_name

    base = load_and_prepare(cfg)
    tagged = tag_movements(base, cfg)

    dist_fp_client, kpi_fp_client, clients_max = kpis_funds_per_client(tagged)
    cpf_total, cpf_month = clients_per_fund(tagged)
    bpi = bpi_stats(tagged)
    comb_freq, comb_resume = combinations(tagged)
    churn_global, churn_fund_month = churn_metrics(tagged)
    acq_global, acq_fund, top_acq_abs, retention = acquisition_and_retention(tagged)
    profiles = client_profiles(tagged)
    flow_detail, flow_matrix = flows_transitions(tagged)

    sheets: Dict[str, pd.DataFrame] = {
        "Base": base,
        "Movements": tagged.drop(columns=["prev_nr_unidades", "next_nr_unidades"]),
        "FundsPerClient": dist_fp_client,
        "KPI_FundsPerClient": kpi_fp_client,
        "Clients_MaxFunds": clients_max,
        "ClientsPerFund": cpf_total,
        "ClientsPerFund_Month": cpf_month,
        "Combinations": comb_freq,
        "Combinations_Resume": comb_resume,
        "Churn_Global_Month": churn_global,
        "Churn_Fund_Month": churn_fund_month,
        "Acquisition_Global_Month": acq_global,
        "Acquisition_Fund_Month": acq_fund,
        "Top_Acquisition_Abs": top_acq_abs,
        "Retention_Fund_Month": retention,
        "Client_Profiles": profiles,
        "Flows_Detail": flow_detail,
        "Flows_Matrix": flow_matrix,
    }
    for k, v in bpi.items():
        sheets[k] = v  # BPI_Stats, BPI_Monthly_Volume, BPI_Extra

    save_all_outputs(sheets, out_excel, csv_dir)
    log.info("✅ LIFE CYCLE pipeline done.")
    log.info("Excel: %s", out_excel)
    log.info("CSVs:  %s", csv_dir)
    return out_excel, csv_dir


# ─────────────────────────────── 3) CLI ───────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="WAR REPORT and LIFE CYCLE pipelines (clean, configurable).")
    parser.add_argument("--log-level", default=os.getenv("LOG_LEVEL", "INFO"), help="Logging level (e.g., INFO, DEBUG)")
    sub = parser.add_subparsers(dest="cmd", required=True)

    # WAR
    p_war = sub.add_parser("war", help="Run WAR REPORT pipeline on daily Excel files.")
    p_war.add_argument("--input", default=os.getenv("WAR_INPUT_DIR", "./data/war_input"), help="Input folder with daily XLSX files.")
    p_war.add_argument("--output", default=os.getenv("WAR_OUTPUT_DIR", "./out/war"), help="Output folder.")
    p_war.add_argument("--start", default=None, help="Start date YYYY-MM-DD (default: 2024-01-02).")
    p_war.add_argument("--end", default=None, help="End date YYYY-MM-DD (default: today).")
    p_war.add_argument("--excel-name", default="Dados_Continuos.xlsx", help="Generic Excel name when period is continuous.")
    p_war.add_argument("--glob", default="*.xlsx", help="Filename pattern.")
    p_war.add_argument("--usecols", default="A:I", help="Excel read usecols.")
    p_war.add_argument("--skiprows", type=int, default=10, help="Excel read skiprows.")
    p_war.add_argument("--evolution-funds", nargs="*", default=None, help="List of fund names for 'Evolution_' sheets.")
    p_war.add_argument("--arima-funds", nargs="*", default=None, help="List of fund names for 'Forecast_' sheets.")
    p_war.add_argument("--arima-horizon", type=int, default=10, help="Forecast horizon (business days).")

    # LIFE
    p_life = sub.add_parser("life", help="Run CLIENT FUNDS LIFE CYCLE pipeline on monthly CSV.")
    p_life.add_argument("--csv", default=os.getenv("LIFE_INPUT_CSV", "./data/life/data.csv"), help="Input CSV path.")
    p_life.add_argument("--output", default=os.getenv("LIFE_OUTPUT_DIR", "./out/life"), help="Output folder.")
    p_life.add_argument("--excel-name", default="Funds_Life.xlsx", help="Output Excel filename.")
    p_life.add_argument("--first-month-known", default="2023-11-30", help="First snapshot month (YYYY-MM-DD) for special rule.")

    args = parser.parse_args()

    # init logging once flags/env are known
    _setup_logging(args.log_level)

    if args.cmd == "war":
        run_war_report(
            input_dir=args.input,
            output_dir=args.output,
            start=args.start,
            end=args.end,
            excel_generic_name=args.excel_name,
            file_glob=args.glob,
            read_usecols=args.usecols,
            read_skiprows=args.skiprows,
            evolution_funds=args.evolution_funds,
            arima_funds=args.arima_funds,
            arima_horizon=args.arima_horizon,
        )
    elif args.cmd == "life":
        run_life_cycle(
            csv_path=args.csv,
            output_dir=args.output,
            excel_name=args.excel_name,
            first_month_known=args.first_month_known,
        )


if __name__ == "__main__":
    main()
